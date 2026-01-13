#!/usr/bin/env python3
"""
Analyze data completeness by comparing Excel source with database
"""

import openpyxl
import json
import subprocess
import sys

# Excel file path
EXCEL_FILE = '/Users/jameslo-aa/ngo_going_out/data/4_NGO going out_RA_project 614.xlsx'

# Field mapping from import_orgs.js
FIELD_MAPPING = {
    '编号': 'id',
    '组织名称': 'org_name',
    '中促会': 'in_cnie',
    '民促会': 'in_cace',
    '联合国': 'in_un',
    '成立时间': 'founded_date',
    '出海时间': 'go_global_date',
    '领导人': 'leaders',
    '重要员工': 'key_staff',
    '资本类型': 'capital_type',
    '注册地': 'reg_location',
    '注册形式': 'reg_type',
    '捐赠金额（出海前）标注年份': 'donation_pre_year',
    '捐赠金额（出海后）': 'donation_post',
    '官网的组织理念': 'mission',
    '组织结构（参考年报）': 'org_structure',
    '是否有独立的海外办公室——组织结构': 'has_overseas_office',
    '官网关于海外项目的组织理念——目标': 'overseas_mission',
    '海外项目的名称': 'overseas_projects',
    '海外涉及的地区': 'overseas_regions',
    '海外服务内容': 'overseas_services',
    '服务形式': 'service_mode',
    '主要成员是否有官方背景': 'has_official_background',
    '主要信息来源': 'sources',
    '是否有网上披露': 'disclosed_online',
    '是否持续性披露': 'disclosed_continuous',
    '走出去程度': 'go_out_level',
}

print("=" * 80)
print("DATA COMPLETENESS ANALYSIS")
print("=" * 80)
print()

# Step 1: Read Excel file
print("Step 1: Reading Excel file...")
print(f"File: {EXCEL_FILE}")
print()

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    total_rows = ws.max_row - 1  # Exclude header

    print(f"✓ Excel loaded successfully")
    print(f"  Total rows: {total_rows}")
    print(f"  Total columns: {len(headers)}")
    print()

except Exception as e:
    print(f"✗ Error reading Excel: {e}")
    sys.exit(1)

# Step 2: Analyze Excel data
print("Step 2: Analyzing Excel data...")
print()

excel_stats = {}
for col_idx, header in enumerate(headers, 1):
    if not header:
        continue

    non_empty = 0
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        value = row[0].value
        if value is not None and str(value).strip() not in ['', '-', 'null']:
            non_empty += 1

    excel_stats[header] = {
        'non_empty': non_empty,
        'empty': total_rows - non_empty,
        'percentage': round(non_empty / total_rows * 100, 1) if total_rows > 0 else 0
    }

print(f"✓ Excel analysis complete")
print()

# Step 3: Get database field statistics
print("Step 3: Querying database statistics...")
print()

# Build SQL to count non-null values for each field
db_fields = [
    'id', 'org_name', 'in_cnie', 'in_cace', 'in_un', 'founded_date', 'go_global_date',
    'leaders', 'key_staff', 'capital_type', 'reg_location', 'reg_type', 'donation_pre',
    'donation_pre_year', 'donation_post', 'mission', 'org_structure', 'has_overseas_office',
    'overseas_mission', 'overseas_projects', 'overseas_regions', 'overseas_services',
    'service_mode', 'has_official_background', 'sources', 'disclosed_online',
    'disclosed_continuous', 'go_out_level', 'logo_url'
]

count_parts = [f"COUNT({field}) as {field}_count" for field in db_fields]
sql = f"SELECT COUNT(*) as total, {', '.join(count_parts)} FROM orgs"

try:
    result = subprocess.run(
        ['npx', 'wrangler', 'd1', 'execute', 'ngo_going_out', '--remote', '--command', sql, '--json'],
        capture_output=True,
        text=True,
        check=True
    )

    # Parse JSON output
    output = json.loads(result.stdout)
    db_stats_raw = output[0]['results'][0]
    total_db_records = db_stats_raw['total']

    db_stats = {}
    for field in db_fields:
        count = db_stats_raw.get(f'{field}_count', 0)
        db_stats[field] = {
            'non_empty': count,
            'empty': total_db_records - count,
            'percentage': round(count / total_db_records * 100, 1) if total_db_records > 0 else 0
        }

    print(f"✓ Database query complete")
    print(f"  Total records: {total_db_records}")
    print()

except Exception as e:
    print(f"✗ Error querying database: {e}")
    sys.exit(1)

# Step 4: Compare and generate report
print("=" * 80)
print("COMPARISON REPORT")
print("=" * 80)
print()

print(f"Excel Records: {total_rows}")
print(f"Database Records: {total_db_records}")
print()

if total_rows != total_db_records:
    print(f"⚠️  WARNING: Record count mismatch! Difference: {abs(total_rows - total_db_records)}")
    print()

# Compare fields
print("-" * 80)
print(f"{'Excel Column':<45} {'DB Field':<20} {'Excel':<10} {'DB':<10} {'Diff':<10}")
print("-" * 80)

issues = []
for excel_col, db_field in FIELD_MAPPING.items():
    if excel_col not in excel_stats:
        issues.append(f"Excel column '{excel_col}' not found in Excel file")
        continue

    if db_field not in db_stats:
        issues.append(f"Database field '{db_field}' not found in database")
        print(f"{excel_col:<45} {db_field:<20} {excel_stats[excel_col]['non_empty']:<10} {'N/A':<10} {'MISSING':<10}")
        continue

    excel_count = excel_stats[excel_col]['non_empty']
    db_count = db_stats[db_field]['non_empty']
    diff = db_count - excel_count

    status = ""
    if diff < 0:
        status = f"⚠️ -{abs(diff)}"
        issues.append(f"{db_field}: Database has {abs(diff)} fewer non-null values than Excel")
    elif diff > 0:
        status = f"✓ +{diff}"
    else:
        status = "✓"

    print(f"{excel_col:<45} {db_field:<20} {excel_count:<10} {db_count:<10} {status:<10}")

print("-" * 80)
print()

# Check for unmapped Excel columns
print("Excel columns NOT in field mapping:")
unmapped_excel = [col for col in headers if col and col not in FIELD_MAPPING]
if unmapped_excel:
    for col in unmapped_excel:
        print(f"  - {col} (has {excel_stats.get(col, {}).get('non_empty', 0)} non-empty values)")
else:
    print("  (none)")
print()

# Check for unmapped database fields
print("Database fields NOT in field mapping:")
unmapped_db = [field for field in db_fields if field not in FIELD_MAPPING.values() and field != 'id']
if unmapped_db:
    for field in unmapped_db:
        print(f"  - {field} (has {db_stats[field]['non_empty']} non-empty values)")
else:
    print("  (none)")
print()

# Summary
print("=" * 80)
print("ISSUES SUMMARY")
print("=" * 80)
if issues:
    for i, issue in enumerate(issues, 1):
        print(f"{i}. {issue}")
else:
    print("✓ No issues found - data appears complete!")
print()

print("Analysis complete!")

