#!/usr/bin/env python3
"""
数据清洗脚本 - 从Excel文件读取并生成清洗后的CSV文件
处理 "4_NGO going out_RA_project 614.xlsx" 文件
"""
import csv
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip3 install openpyxl")
    exit(1)

# 文件路径
INPUT_EXCEL = Path("../data/4_NGO going out_RA_project 614.xlsx")
OUTPUT_ORGS = Path("../data/orgs_clean.csv")
OUTPUT_POLICIES = Path("../data/policies_clean.csv")

def clean_value(val):
    """清理单元格值"""
    if val is None:
        return ""
    if isinstance(val, str):
        val = val.strip()
        # 处理各种空值表示
        if val in ("——", "-", "NA", "N/A", "na", "n/a", ""):
            return ""
        return val
    return val

def clean_number(val):
    """清理数字字段，去除货币符号和逗号"""
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val = val.strip()
        if val in ("——", "-", "NA", "N/A", "na", "n/a", ""):
            return ""
        # 去除货币符号和逗号
        val = val.replace("￥", "").replace(",", "").strip()
        try:
            return float(val)
        except ValueError:
            return ""
    return ""

def clean_date(val):
    """清理日期字段，统一格式为 YYYY-MM-DD 或 YYYY-MM 或 YYYY"""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        if val in ("——", "-", "NA", "N/A", "na", "n/a", ""):
            return ""
        # 保持原样，可能是 "2009年" 或 "1985——4——1" 等格式
        return val
    return str(val)

def normalize_bool(val):
    """标准化布尔值"""
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return 1 if val == 1 else 0 if val == 0 else ""
    s = str(val).strip().lower()
    if s in ("——", "-", "na", "n/a"):
        return ""
    if s in ("1", "1.0", "true", "yes", "是", "有", "y"):
        return 1
    if s in ("0", "0.0", "false", "no", "否", "无", "n"):
        return 0
    return ""

def write_csv(path, fieldnames, rows):
    """写入CSV文件"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

def process_policies():
    """处理政策数据"""
    print("处理政策数据...")
    wb = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
    ws = wb["相关政策收集"]

    # Row 1 是说明文字，Row 2 是表头，数据从 Row 3 开始
    headers = [cell.value for cell in ws[2]]

    required_fields = ["id", "published_date", "title", "doc_type",
                      "issuer_1", "issuer_2", "issuer_3", "issuer_4", "link"]

    # 列索引映射
    col_map = {
        "No.": "id",
        "发布时期": "published_date",
        "题目": "title",
        "属性": "doc_type",
        "发布单位（部委）1": "issuer_1",
        "发布单位（部委）2": "issuer_2",
        "发布单位（部委）3": "issuer_3",
        "发布单位（部委）4": "issuer_4",
        "链接": "link"
    }

    # 创建列索引
    col_indices = {}
    for i, h in enumerate(headers):
        if h in col_map:
            col_indices[col_map[h]] = i

    out_rows = []
    for row_idx in range(3, ws.max_row + 1):  # 从第3行开始
        row = ws[row_idx]
        out = {k: "" for k in required_fields}

        # 提取数据
        for field, col_idx in col_indices.items():
            val = row[col_idx].value
            if field == "id":
                out[field] = int(val) if val else row_idx - 2
            elif field == "published_date":
                out[field] = clean_date(val)
            else:
                out[field] = clean_value(val)

        # 跳过完全空白的行
        if all(str(v).strip() == "" for v in out.values() if v != out["id"]):
            continue

        out_rows.append(out)

    write_csv(OUTPUT_POLICIES, required_fields, out_rows)
    print(f"  完成: {len(out_rows)} 条政策记录 -> {OUTPUT_POLICIES}")

def process_orgs():
    """处理组织数据"""
    print("处理组织数据...")
    wb = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
    ws = wb["组织相关信息"]

    # Row 1 是表头，数据从 Row 2 开始
    headers = [cell.value for cell in ws[1]]

    required_fields = [
        "id", "org_name", "in_cnie", "in_cace", "in_un",
        "founded_date", "go_global_date", "leaders", "key_staff",
        "capital_type", "reg_location", "reg_type",
        "donation_pre", "donation_pre_year", "donation_post", "donation_post_year",
        "mission", "org_structure", "has_overseas_office",
        "overseas_mission", "overseas_projects", "overseas_regions",
        "overseas_services", "service_mode", "has_official_background",
        "sources", "disclosed_online", "disclosed_continuous",
        "go_out_level", "logo_url"
    ]

    # 列索引映射 (注意: 列13是金额，列14是年份)
    col_map = {
        "编号": ("id", 0),
        "组织名称": ("org_name", 1),
        "中促会": ("in_cnie", 2),
        "民促会": ("in_cace", 3),
        "联合国": ("in_un", 4),
        "成立时间": ("founded_date", 5),
        "出海时间": ("go_global_date", 6),
        "领导人": ("leaders", 7),
        "重要员工": ("key_staff", 8),
        "资本类型": ("capital_type", 9),
        "注册地": ("reg_location", 10),
        "注册形式": ("reg_type", 11),
        "捐赠金额（出海前）标注年份": ("donation_pre", 12),  # 实际是金额
        None: [("donation_pre_year", 13), ("donation_post_year", 15)],  # 年份列
        "捐赠金额（出海后）": ("donation_post", 14),
        "官网的组织理念": ("mission", 16),
        "组织结构（参考年报）": ("org_structure", 17),
        "是否有独立的海外办公室——组织结构": ("has_overseas_office", 18),
        "官网关于海外项目的组织理念——目标": ("overseas_mission", 19),
        "海外项目的名称": ("overseas_projects", 20),
        "海外涉及的地区": ("overseas_regions", 21),
        "海外服务内容": ("overseas_services", 22),
        "服务形式": ("service_mode", 23),
        "主要成员是否有官方背景": ("has_official_background", 24),
        "主要信息来源": ("sources", 25),
        "是否有网上披露": ("disclosed_online", 26),
        "是否持续性披露": ("disclosed_continuous", 27),
        "走出去程度": ("go_out_level", 28),
        "官网LOGO或图片": ("logo_url", 29)
    }

    out_rows = []
    for row_idx in range(2, ws.max_row + 1):  # 从第2行开始
        row = ws[row_idx]
        out = {k: "" for k in required_fields}

        # 基本字段
        out["id"] = int(row[0].value) if row[0].value else row_idx - 1
        out["org_name"] = clean_value(row[1].value)

        # 布尔字段
        out["in_cnie"] = normalize_bool(row[2].value)
        out["in_cace"] = normalize_bool(row[3].value)
        out["in_un"] = normalize_bool(row[4].value)
        out["has_overseas_office"] = normalize_bool(row[18].value)
        out["has_official_background"] = normalize_bool(row[24].value)
        out["disclosed_online"] = normalize_bool(row[26].value)
        out["disclosed_continuous"] = normalize_bool(row[27].value)

        # 日期字段
        out["founded_date"] = clean_date(row[5].value)
        out["go_global_date"] = clean_date(row[6].value)

        # 文本字段
        out["leaders"] = clean_value(row[7].value)
        out["key_staff"] = clean_value(row[8].value)
        out["capital_type"] = clean_value(row[9].value)
        out["reg_location"] = clean_value(row[10].value)
        out["reg_type"] = clean_value(row[11].value)
        out["mission"] = clean_value(row[16].value)
        out["org_structure"] = clean_value(row[17].value)
        out["overseas_mission"] = clean_value(row[19].value)
        out["overseas_projects"] = clean_value(row[20].value)
        out["overseas_regions"] = clean_value(row[21].value)
        out["overseas_services"] = clean_value(row[22].value)
        out["service_mode"] = clean_value(row[23].value)
        out["sources"] = clean_value(row[25].value)
        out["go_out_level"] = clean_value(row[28].value)

        # 捐赠金额和年份 (注意列的实际含义)
        out["donation_pre"] = clean_number(row[12].value)
        out["donation_pre_year"] = clean_value(row[13].value)
        out["donation_post"] = clean_number(row[14].value)
        out["donation_post_year"] = clean_value(row[15].value)

        # Logo URL - DISPIMG公式无法使用，设为空
        logo_val = row[29].value
        if logo_val and isinstance(logo_val, str) and not logo_val.startswith("=DISPIMG"):
            out["logo_url"] = clean_value(logo_val)
        else:
            out["logo_url"] = ""

        # 跳过完全空白的行
        if not out["org_name"]:
            continue

        out_rows.append(out)

    write_csv(OUTPUT_ORGS, required_fields, out_rows)
    print(f"  完成: {len(out_rows)} 条组织记录 -> {OUTPUT_ORGS}")

if __name__ == "__main__":
    if not INPUT_EXCEL.exists():
        print(f"错误: 找不到文件 {INPUT_EXCEL}")
        exit(1)

    process_policies()
    process_orgs()
    print("\n数据清洗完成!")
