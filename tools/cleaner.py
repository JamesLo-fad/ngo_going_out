#!/usr/bin/env python3
"""
数据清洗脚本 - 从Excel文件读取并生成清洗后的CSV文件
处理 "4_NGO going out_RA_project 614.xlsx" 文件
输出CSV使用中文列名
"""
import csv
import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip3 install openpyxl")
    exit(1)

# 文件路径
INPUT_EXCEL = Path("../data/4_NGO going out_RA_project 614.xlsx")
OUTPUT_ORGS = Path("../data/orgs_clean.csv")
OUTPUT_POLICIES = Path("../data/policies_clean.csv")

def clean_value(val):
    """清理单元格值 - 处理各种空值表示"""
    if val is None:
        return ""
    if isinstance(val, str):
        val = val.strip()
        # 处理各种空值表示
        empty_values = ["——", "-", "--", "---", "————",
                       "NA", "N/A", "na", "n/a", "n.a.", "N.A.",
                       "无", "暂无", "未知", "/", "\\"]
        if val in empty_values or val == "":
            return ""
        return val
    return val

def clean_number(val):
    """清理数字字段，去除货币符号、逗号，处理中文数字单位"""
    if val is None or val == "":
        return ""

    # 如果已经是数字，直接返回
    if isinstance(val, (int, float)):
        return val

    if isinstance(val, str):
        val = val.strip()
        # 检查空值
        empty_values = ["——", "-", "NA", "N/A", "na", "n/a", "无", "暂无", "未知", "未披露"]
        if val in empty_values or val == "":
            return ""

        # 去除货币符号和逗号
        val = val.replace("￥", "").replace("¥", "").replace(",", "").replace("，", "").strip()

        # 处理中文数字单位
        if "万" in val or "亿" in val:
            # 提取数字部分
            num_match = re.search(r'([\d.]+)', val)
            if num_match:
                num = float(num_match.group(1))
                if "亿" in val:
                    return num * 100000000
                elif "万" in val:
                    return num * 10000

        # 尝试转换为浮点数
        try:
            return float(val)
        except ValueError:
            # 如果包含非数字文本（如"约"、"超过"等），返回空
            return ""

    return ""

def clean_date(val):
    """
    清理日期字段，统一格式
    - datetime对象 → YYYY-MM-DD
    - "2009年" → "2009"
    - "1985——4——1" → "1985-04-01"
    - 保持其他格式但做基本清理
    """
    if val is None or val == "":
        return ""

    # datetime对象直接格式化
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    # 字符串处理
    if isinstance(val, str):
        val = val.strip()

        # 检查空值
        empty_values = ["——", "-", "NA", "N/A", "na", "n/a", "无", "暂无", "未知"]
        if val in empty_values or val == "":
            return ""

        # 处理 "2009年" 格式 → "2009"
        if val.endswith("年"):
            year = val[:-1].strip()
            # 验证是否为有效年份
            try:
                if len(year) == 4 and year.isdigit():
                    return year
            except:
                pass

        # 处理 "1985——4——1" 格式 → "1985-04-01"
        if "——" in val:
            parts = val.split("——")
            if len(parts) == 3:
                try:
                    year, month, day = parts
                    year = year.strip()
                    month = month.strip().zfill(2)
                    day = day.strip().zfill(2)
                    # 验证日期有效性
                    datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
                    return f"{year}-{month}-{day}"
                except:
                    pass  # 如果转换失败，继续处理

        # 处理 "2009-01" 或 "2009/01" 格式
        if re.match(r'^\d{4}[-/]\d{1,2}$', val):
            parts = re.split(r'[-/]', val)
            return f"{parts[0]}-{parts[1].zfill(2)}"

        # 处理 "2009-01-15" 或 "2009/01/15" 格式
        if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$', val):
            parts = re.split(r'[-/]', val)
            try:
                return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
            except:
                pass

        # 其他格式保持原样
        return val

    return str(val)

def normalize_bool(val):
    """标准化布尔值为 1/0 或空字符串"""
    if val is None or val == "":
        return ""

    # 如果是数字
    if isinstance(val, (int, float)):
        if val == 1 or val == 1.0:
            return 1
        elif val == 0 or val == 0.0:
            return 0
        else:
            return ""

    # 字符串处理
    s = str(val).strip().lower()

    # 空值
    if s in ("——", "-", "na", "n/a", "无", "暂无", "未知", ""):
        return ""

    # 真值
    if s in ("1", "1.0", "true", "yes", "是", "有", "y"):
        return 1

    # 假值
    if s in ("0", "0.0", "false", "no", "否", "无", "n"):
        return 0

    return ""

def write_csv(path, fieldnames, rows):
    """写入CSV文件"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

def process_policies():
    """处理政策数据"""
    print("处理政策数据...")
    wb = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
    ws = wb["相关政策收集"]

    # Row 1 是说明文字，Row 2 是表头，数据从 Row 3 开始
    # 使用中文列名
    required_fields = ["编号", "发布时期", "题目", "属性",
                      "发布单位（部委）1", "发布单位（部委）2",
                      "发布单位（部委）3", "发布单位（部委）4", "链接"]

    out_rows = []
    for row_idx in range(3, ws.max_row + 1):  # 从第3行开始
        row = ws[row_idx]
        out = {k: "" for k in required_fields}

        # 提取数据
        out["编号"] = int(row[0].value) if row[0].value else row_idx - 2
        out["发布时期"] = clean_date(row[1].value)
        out["题目"] = clean_value(row[2].value)
        out["属性"] = clean_value(row[3].value)
        out["发布单位（部委）1"] = clean_value(row[4].value)
        out["发布单位（部委）2"] = clean_value(row[5].value)
        out["发布单位（部委）3"] = clean_value(row[6].value)
        out["发布单位（部委）4"] = clean_value(row[7].value)
        out["链接"] = clean_value(row[8].value)

        # 跳过完全空白的行
        if all(str(v).strip() == "" for k, v in out.items() if k != "编号"):
            continue

        out_rows.append(out)

    write_csv(OUTPUT_POLICIES, required_fields, out_rows)
    print(f"  完成: {len(out_rows)} 条政策记录 -> {OUTPUT_POLICIES}")

def process_orgs():
    """处理组织数据"""
    print("处理组织数据...")
    wb = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
    ws = wb["组织相关信息"]

    # Row 1 是表头，数据从 Row 2 开始
    # 使用中文列名（更清晰的命名）
    required_fields = [
        "编号", "组织名称", "中促会", "民促会", "联合国",
        "成立时间", "出海时间", "领导人", "重要员工",
        "资本类型", "注册地", "注册形式",
        "捐赠金额（出海前）", "捐赠年份（出海前）",
        "捐赠金额（出海后）", "捐赠年份（出海后）",
        "官网的组织理念", "组织结构（参考年报）",
        "是否有独立的海外办公室——组织结构",
        "官网关于海外项目的组织理念——目标",
        "海外项目的名称", "海外涉及的地区",
        "海外服务内容", "服务形式",
        "主要成员是否有官方背景", "主要信息来源",
        "是否有网上披露", "是否持续性披露",
        "走出去程度", "官网LOGO或图片"
    ]

    out_rows = []
    for row_idx in range(2, ws.max_row + 1):  # 从第2行开始
        row = ws[row_idx]
        out = {k: "" for k in required_fields}

        # 基本字段
        out["编号"] = int(row[0].value) if row[0].value else row_idx - 1
        out["组织名称"] = clean_value(row[1].value)

        # 布尔字段
        out["中促会"] = normalize_bool(row[2].value)
        out["民促会"] = normalize_bool(row[3].value)
        out["联合国"] = normalize_bool(row[4].value)
        out["是否有独立的海外办公室——组织结构"] = normalize_bool(row[18].value)
        out["主要成员是否有官方背景"] = normalize_bool(row[24].value)
        out["是否有网上披露"] = normalize_bool(row[26].value)
        out["是否持续性披露"] = normalize_bool(row[27].value)

        # 日期字段
        out["成立时间"] = clean_date(row[5].value)
        out["出海时间"] = clean_date(row[6].value)

        # 文本字段
        out["领导人"] = clean_value(row[7].value)
        out["重要员工"] = clean_value(row[8].value)
        out["资本类型"] = clean_value(row[9].value)
        out["注册地"] = clean_value(row[10].value)
        out["注册形式"] = clean_value(row[11].value)
        out["官网的组织理念"] = clean_value(row[16].value)
        out["组织结构（参考年报）"] = clean_value(row[17].value)
        out["官网关于海外项目的组织理念——目标"] = clean_value(row[19].value)
        out["海外项目的名称"] = clean_value(row[20].value)
        out["海外涉及的地区"] = clean_value(row[21].value)
        out["海外服务内容"] = clean_value(row[22].value)
        out["服务形式"] = clean_value(row[23].value)
        out["主要信息来源"] = clean_value(row[25].value)
        out["走出去程度"] = clean_value(row[28].value)

        # 捐赠金额和年份 (注意列的实际含义)
        out["捐赠金额（出海前）"] = clean_number(row[12].value)
        out["捐赠年份（出海前）"] = clean_date(row[13].value)
        out["捐赠金额（出海后）"] = clean_number(row[14].value)
        out["捐赠年份（出海后）"] = clean_date(row[15].value)

        # Logo URL - DISPIMG公式无法使用，设为空
        logo_val = row[29].value
        if logo_val and isinstance(logo_val, str) and not logo_val.startswith("=DISPIMG"):
            out["官网LOGO或图片"] = clean_value(logo_val)
        else:
            out["官网LOGO或图片"] = ""

        # 跳过完全空白的行
        if not out["组织名称"]:
            continue

        out_rows.append(out)

    write_csv(OUTPUT_ORGS, required_fields, out_rows)
    print(f"  完成: {len(out_rows)} 条组织记录 -> {OUTPUT_ORGS}")

if __name__ == "__main__":
    if not INPUT_EXCEL.exists():
        print(f"错误: 找不到文件 {INPUT_EXCEL}")
        exit(1)

    process_policies()
    process_orgs()
    print("\n数据清洗完成!")
    print("\n注意: CSV文件使用中文列名")
    print("导入脚本会自动将中文列名映射到数据库的英文字段名")
