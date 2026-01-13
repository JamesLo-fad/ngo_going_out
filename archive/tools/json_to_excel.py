#!/usr/bin/env python3
"""
Convert database JSON export to Excel file
Reads the JSON export and creates a properly formatted Excel file
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Configuration
JSON_FILE = '/tmp/db_export.json'
OUTPUT_FILE = f'/Users/jameslo-aa/ngo_going_out/data/ngo_database_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

# Column mapping: database field -> Chinese header
COLUMN_MAPPING = {
    'id': '编号',
    'org_name': '组织名称',
    'in_cnie': '中促会',
    'in_cace': '民促会',
    'in_un': '联合国',
    'founded_date': '成立时间',
    'go_global_date': '出海时间',
    'leaders': '领导人',
    'key_staff': '重要员工',
    'capital_type': '资本类型',
    'reg_location': '注册地',
    'reg_type': '注册形式',
    'donation_pre': '捐赠金额（出海前）',
    'donation_pre_year': '捐赠年份（出海前）',
    'donation_post': '捐赠金额（出海后）',
    'mission': '官网的组织理念',
    'org_structure': '组织结构（参考年报）',
    'has_overseas_office': '是否有独立的海外办公室——组织结构',
    'overseas_mission': '官网关于海外项目的组织理念——目标',
    'overseas_projects': '海外项目的名称',
    'overseas_regions': '海外涉及的地区',
    'overseas_services': '海外服务内容',
    'service_mode': '服务形式',
    'has_official_background': '主要成员是否有官方背景',
    'sources': '主要信息来源',
    'disclosed_online': '是否有网上披露',
    'disclosed_continuous': '是否持续性披露',
    'go_out_level': '走出去程度',
    'logo_url': '官网LOGO或图片'
}

def format_value(value, field_name):
    """Format value for Excel output"""
    if value is None:
        return ''

    # Boolean fields: convert 1/0 to 是/否
    if field_name in ['in_cnie', 'in_cace', 'in_un', 'has_overseas_office',
                      'has_official_background', 'disclosed_online', 'disclosed_continuous']:
        if value == 1:
            return '是'
        elif value == 0:
            return '否'
        else:
            return ''

    return value

def main():
    try:
        print(f'\n📊 导出数据库到 Excel')
        print(f'   输入: {JSON_FILE}')
        print(f'   输出: {OUTPUT_FILE}\n')

        # Read JSON data
        print('📥 读取数据...')
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)

        records = data[0]['results']
        print(f'   ✓ 获取 {len(records)} 条记录\n')

        # Create workbook
        print('📝 创建 Excel 文件...')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'NGO Organizations'

        # Get column names in order
        if not records:
            raise Exception("No records to export")

        db_columns = list(records[0].keys())

        # Write headers with Chinese names
        headers = [COLUMN_MAPPING.get(col, col) for col in db_columns]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for record in records:
            row_data = [format_value(record[col], col) for col in db_columns]
            ws.append(row_data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save workbook
        wb.save(OUTPUT_FILE)

        print(f'   ✓ 已写入 {len(records)} 条记录\n')

        print('✅ 导出完成!')
        print(f'   文件: {OUTPUT_FILE}')
        print(f'   记录数: {len(records)}')

        # Show summary stats
        logo_count = sum(1 for r in records if r.get('logo_url'))
        disclosed_online_count = sum(1 for r in records if r.get('disclosed_online'))
        disclosed_continuous_count = sum(1 for r in records if r.get('disclosed_continuous'))
        go_out_level_count = sum(1 for r in records if r.get('go_out_level'))

        print(f'\n📈 数据统计:')
        print(f'   Logo URLs: {logo_count}')
        print(f'   网上披露: {disclosed_online_count}')
        print(f'   持续性披露: {disclosed_continuous_count}')
        print(f'   走出去程度: {go_out_level_count}')

    except Exception as e:
        print(f'\n❌ 导出失败: {str(e)}')
        import traceback
        traceback.print_exc()
        exit(1)

if __name__ == '__main__':
    main()
