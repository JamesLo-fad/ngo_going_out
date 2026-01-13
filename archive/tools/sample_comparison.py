#!/usr/bin/env python3
"""
Compare sample records between Excel and database
"""

import openpyxl
import json
import subprocess

EXCEL_FILE = '/Users/jameslo-aa/ngo_going_out/data/4_NGO going out_RA_project 614.xlsx'

# Sample IDs to check
SAMPLE_IDS = [1, 10, 50, 100, 200]

print("=" * 80)
print("SAMPLE RECORD COMPARISON")
print("=" * 80)
print()

# Load Excel
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Get Excel data for sample IDs
excel_data = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    org_id = row[0]  # First column is ID
    if org_id in SAMPLE_IDS:
        excel_data[org_id] = dict(zip(headers, row))

# Get database data for sample IDs
db_data = {}
for org_id in SAMPLE_IDS:
    result = subprocess.run(
        ['npx', 'wrangler', 'd1', 'execute', 'ngo_going_out', '--remote',
         '--command', f'SELECT * FROM orgs WHERE id = {org_id}', '--json'],
        capture_output=True,
        text=True,
        check=True
    )
    output = json.loads(result.stdout)
    if output[0]['results']:
        db_data[org_id] = output[0]['results'][0]

# Compare specific problematic fields
problematic_fields = [
    ('资本类型', 'capital_type'),
    ('出海时间', 'go_global_date'),
    ('重要员工', 'key_staff'),
    ('捐赠金额（出海前）标注年份', 'donation_pre_year'),
    ('是否有网上披露', 'disclosed_online'),
    ('是否持续性披露', 'disclosed_continuous'),
    ('走出去程度', 'go_out_level'),
]

for org_id in SAMPLE_IDS:
    if org_id not in excel_data or org_id not in db_data:
        continue

    print(f"Organization ID: {org_id}")
    print(f"Name: {excel_data[org_id].get('组织名称', 'N/A')}")
    print("-" * 80)

    for excel_col, db_field in problematic_fields:
        excel_val = excel_data[org_id].get(excel_col)
        db_val = db_data[org_id].get(db_field)

        # Show raw values
        excel_repr = repr(excel_val) if excel_val is not None else 'None'
        db_repr = repr(db_val) if db_val is not None else 'None'

        match = "✓" if excel_val == db_val else "✗"

        print(f"{excel_col} ({db_field}):")
        print(f"  Excel: {excel_repr}")
        print(f"  DB:    {db_repr}")
        print(f"  Match: {match}")
        print()

    print()

print("=" * 80)
print("Analysis of cleanValue() function behavior:")
print("=" * 80)
print()
print("The cleanValue() function converts these to NULL:")
print("  - Empty string: ''")
print("  - Dash: '-'")
print("  - String 'null' (case-insensitive)")
print()
print("This is likely causing the data loss!")
print()
