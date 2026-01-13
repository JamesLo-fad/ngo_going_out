#!/usr/bin/env python3
"""
Check what values exist in Excel for fields that are all NULL in database
"""

import openpyxl

EXCEL_FILE = '/Users/jameslo-aa/ngo_going_out/data/4_NGO going out_RA_project 614.xlsx'

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Find column indices
disclosed_online_idx = headers.index('是否有网上披露') + 1
disclosed_continuous_idx = headers.index('是否持续性披露') + 1
go_out_level_idx = headers.index('走出去程度') + 1

print("=" * 80)
print("VALUES IN EXCEL FOR FIELDS THAT ARE ALL NULL IN DATABASE")
print("=" * 80)
print()

# Collect unique values
disclosed_online_values = set()
disclosed_continuous_values = set()
go_out_level_values = set()

for row in ws.iter_rows(min_row=2, max_row=20):  # Check first 20 rows
    disclosed_online_values.add(repr(row[disclosed_online_idx - 1].value))
    disclosed_continuous_values.add(repr(row[disclosed_continuous_idx - 1].value))
    go_out_level_values.add(repr(row[go_out_level_idx - 1].value))

print("是否有网上披露 (disclosed_online) - unique values in first 20 rows:")
for val in sorted(disclosed_online_values):
    print(f"  {val}")
print()

print("是否持续性披露 (disclosed_continuous) - unique values in first 20 rows:")
for val in sorted(disclosed_continuous_values):
    print(f"  {val}")
print()

print("走出去程度 (go_out_level) - unique values in first 20 rows:")
for val in sorted(go_out_level_values):
    print(f"  {val}")
print()

# Check if these are being treated as "是/否" but import script expects 1/0
print("=" * 80)
print("DIAGNOSIS")
print("=" * 80)
print()
print("These fields contain '是' and '否' in Excel, but the import script")
print("likely expects 1/0 or TRUE/FALSE for INTEGER fields.")
print()
print("The cleanValue() function doesn't convert '是'/'否' to 1/0,")
print("so they get inserted as text into INTEGER fields, which SQLite")
print("converts to 0 or NULL.")
print()
