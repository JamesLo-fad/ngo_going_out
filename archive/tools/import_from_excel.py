#!/usr/bin/env python3
"""
Import organizations directly from Excel file to D1 database
This preserves Chinese dash '——' values that are lost in CSV conversion
"""

import openpyxl
import subprocess
import json
import sys
import os

EXCEL_FILE = '/Users/jameslo-aa/ngo_going_out/data/4_NGO going out_RA_project 614.xlsx'
DB_NAME = os.environ.get('D1_DB_NAME')

if not DB_NAME:
    print('❌ 错误: 请设置 D1_DB_NAME 环境变量')
    print('   例如: export D1_DB_NAME=ngo_going_out')
    sys.exit(1)

def parse_yes_no(val):
    """Convert Chinese yes/no to 1/0"""
    if not val:
        return None
    s = str(val).strip().lower()
    if s in ['是', 'yes', '1', 'true']:
        return 1
    if s in ['否', 'no', '0', 'false']:
        return 0
    return None

def clean_value(val):
    """Clean value but preserve Chinese dash '——'"""
    if val is None:
        return None
    s = str(val).strip()
    # Only convert single dash '-' to null, not Chinese dash '——'
    if s == '' or s == '-' or s.lower() == 'null':
        return None
    return s

def parse_float(val):
    """Parse float value"""
    if not val:
        return None
    try:
        return float(val)
    except:
        return None

def esc(v):
    """Escape value for SQL"""
    if v is None or v == '':
        return 'NULL'
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str) and v.isdigit():
        return v
    return "'" + str(v).replace("'", "''") + "'"

def d1_exec(sql):
    """Execute SQL command on D1 database"""
    result = subprocess.run(
        ['npx', 'wrangler', 'd1', 'execute', DB_NAME, '--remote', '--command', sql],
        capture_output=True,
        text=True
    )
    if result.returncode != 0:
        raise Exception(f"D1 execute failed: {result.stderr}")
    return result.stdout

print(f'\n📊 从 Excel 导入组织数据')
print(f'   数据库: {DB_NAME}')
print(f'   文件: {EXCEL_FILE}\n')

# Load Excel file
print('📥 读取 Excel 文件...')
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active
headers = [cell.value for cell in ws[1]]
print(f'   找到 {len(headers)} 列\n')

# Create column index mapping
col_map = {header: idx for idx, header in enumerate(headers)}

# Get column indices
def get_col(row, name):
    """Get cell value by column name"""
    idx = col_map.get(name)
    if idx is None:
        return None
    return row[idx].value

print('🔄 开始导入数据...\n')

total = 0
success = 0
errors = 0

for row in ws.iter_rows(min_row=2):
    total += 1
    try:
        # Extract values from Excel
        org_id = int(get_col(row, '编号') or total)
        org_name = clean_value(get_col(row, '组织名称'))

        if not org_name:
            continue

        # Build SQL INSERT statement
        sql = f"""
        INSERT INTO orgs (
          id, org_name, in_cnie, in_cace, in_un, founded_date, go_global_date,
          leaders, key_staff, capital_type, reg_location, reg_type,
          donation_pre, donation_pre_year, donation_post,
          mission, org_structure, has_overseas_office, overseas_mission,
          overseas_projects, overseas_regions, overseas_services, service_mode,
          has_official_background, sources, disclosed_online, disclosed_continuous,
          go_out_level, logo_url
        ) VALUES (
          {esc(org_id)}, {esc(org_name)},
          {esc(parse_yes_no(get_col(row, '中促会')))},
          {esc(parse_yes_no(get_col(row, '民促会')))},
          {esc(parse_yes_no(get_col(row, '联合国')))},
          {esc(clean_value(get_col(row, '成立时间')))},
          {esc(clean_value(get_col(row, '出海时间')))},
          {esc(clean_value(get_col(row, '领导人')))},
          {esc(clean_value(get_col(row, '重要员工')))},
          {esc(clean_value(get_col(row, '资本类型')))},
          {esc(clean_value(get_col(row, '注册地')))},
          {esc(clean_value(get_col(row, '注册形式')))},
          {esc(parse_float(get_col(row, '捐赠金额（出海前）')))},
          {esc(clean_value(get_col(row, '捐赠金额（出海前）标注年份')))},
          {esc(parse_float(get_col(row, '捐赠金额（出海后）')))},
          {esc(clean_value(get_col(row, '官网的组织理念')))},
          {esc(clean_value(get_col(row, '组织结构（参考年报）')))},
          {esc(parse_yes_no(get_col(row, '是否有独立的海外办公室——组织结构')))},
          {esc(clean_value(get_col(row, '官网关于海外项目的组织理念——目标')))},
          {esc(clean_value(get_col(row, '海外项目的名称')))},
          {esc(clean_value(get_col(row, '海外涉及的地区')))},
          {esc(clean_value(get_col(row, '海外服务内容')))},
          {esc(clean_value(get_col(row, '服务形式')))},
          {esc(parse_yes_no(get_col(row, '主要成员是否有官方背景')))},
          {esc(clean_value(get_col(row, '主要信息来源')))},
          {esc(parse_yes_no(get_col(row, '是否有网上披露')))},
          {esc(parse_yes_no(get_col(row, '是否持续性披露')))},
          {esc(clean_value(get_col(row, '走出去程度')))},
          NULL
        ) ON CONFLICT(id) DO UPDATE SET
          org_name=excluded.org_name, in_cnie=excluded.in_cnie,
          in_cace=excluded.in_cace, in_un=excluded.in_un,
          founded_date=excluded.founded_date, go_global_date=excluded.go_global_date,
          leaders=excluded.leaders, key_staff=excluded.key_staff,
          capital_type=excluded.capital_type, reg_location=excluded.reg_location,
          reg_type=excluded.reg_type, donation_pre=excluded.donation_pre,
          donation_pre_year=excluded.donation_pre_year, donation_post=excluded.donation_post,
          mission=excluded.mission, org_structure=excluded.org_structure,
          has_overseas_office=excluded.has_overseas_office,
          overseas_mission=excluded.overseas_mission,
          overseas_projects=excluded.overseas_projects,
          overseas_regions=excluded.overseas_regions,
          overseas_services=excluded.overseas_services,
          service_mode=excluded.service_mode,
          has_official_background=excluded.has_official_background,
          sources=excluded.sources, disclosed_online=excluded.disclosed_online,
          disclosed_continuous=excluded.disclosed_continuous,
          go_out_level=excluded.go_out_level,
          logo_url=COALESCE(orgs.logo_url, excluded.logo_url);
        """

        # Execute SQL
        d1_exec(sql)
        success += 1

        if success % 50 == 0:
            print(f'   已导入: {success} 条记录...')

    except Exception as e:
        errors += 1
        print(f'❌ 导入第 {total} 行失败: {str(e)[:100]}')

print(f'\n✅ 导入完成!')
print(f'   成功: {success} 条记录')
if errors > 0:
    print(f'   失败: {errors} 条记录')

