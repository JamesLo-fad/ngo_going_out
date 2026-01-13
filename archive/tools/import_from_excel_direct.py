#!/usr/bin/env python3
"""
Direct Excel-to-D1 import script that bypasses CSV to avoid newline parsing bugs.
This script handles multi-line cells, commas, and special characters correctly.
"""

import openpyxl
import json
import re
import subprocess
import sys

EXCEL_FILE = '/Users/jameslo-aa/ngo_going_out/data/4_NGO going out_RA_project 614.xlsx'
LOGO_BACKUP = '/tmp/logo_backup.json'
DB_NAME = 'ngo_going_out'

def normalize_date(val):
    """Normalize date to Chinese format"""
    if not val or val in ['——', '-', 'null', None]:
        return '——'

    s = str(val).strip()

    # Already in correct format
    if re.match(r'^\d{4}\s*年(\s*\d{1,2}\s*月)?(\s*\d{1,2}\s*日)?$', s):
        return s

    # Format: "1985——4——1" (Chinese dash)
    match = re.match(r'^(\d{4})[——\-]+(\d{1,2})[——\-]+(\d{1,2})$', s)
    if match:
        year, month, day = match.groups()
        return f'{year} 年 {int(month)} 月 {int(day)} 日'

    # Format: "2009年" (year only)
    match = re.match(r'^(\d{4})\s*年$', s)
    if match:
        return f'{match.group(1)} 年'

    # Format: "1905-07-10" or "1905-07-10 00:00:00"
    match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})(\s+\d{2}:\d{2}:\d{2})?$', s)
    if match:
        year, month, day = match.groups()[:3]
        return f'{year} 年 {int(month)} 月 {int(day)} 日'

    # Remove parenthetical explanations
    s = re.sub(r'[（(].*?[）)]', '', s).strip()
    if s and re.match(r'^\d{4}$', s):
        return f'{s} 年'

    return s if s else '——'

def parse_yes_no(val):
    """Convert to 1/0 for database"""
    if val is None or val == '':
        return None
    s = str(val).strip().lower()
    if s in ['1', '1.0', 'true', '是']:
        return 1
    if s in ['0', '0.0', 'false', '否']:
        return 0
    return None

def clean_value(val):
    """Clean text value - handles multi-line cells correctly"""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() == 'null' or s == '':
        return None
    # Keep newlines - they are legitimate content in multi-line cells
    # Just normalize them to ensure consistency
    s = s.replace('\r\n', '\n').replace('\r', '\n')
    return s

def load_logo_backup():
    """Load logo URLs from backup"""
    try:
        with open(LOGO_BACKUP, 'r') as f:
            data = json.load(f)
            backup = {}
            for row in data[0]['results']:
                backup[row['id']] = row['logo_url']
            return backup
    except Exception as e:
        print(f'Warning: Could not load logo backup: {e}')
        return {}

def execute_sql(db_name, sql):
    """Execute SQL command via wrangler"""
    cmd = ['npx', 'wrangler', 'd1', 'execute', db_name, '--remote', '--command', sql]
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        raise Exception(f"SQL execution failed: {result.stderr}")

    return result.stdout

def escape_sql_string(s):
    """Escape string for SQL insertion"""
    if s is None:
        return 'NULL'
    # Escape single quotes by doubling them
    escaped = str(s).replace("'", "''")
    return f"'{escaped}'"

def main():
    print('='*100)
    print('📥 Direct Excel-to-D1 Import (Bypassing CSV)')
    print('='*100)

    # Load Excel
    print('\n1. Loading Excel file...')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    print(f'   ✓ Loaded: {ws.max_row} rows, {ws.max_column} columns')

    # Load logo backup
    print('\n2. Loading logo backup...')
    logo_backup = load_logo_backup()
    print(f'   ✓ Loaded {len(logo_backup)} logo URLs')

    # Column mapping (0-indexed Excel columns to field names)
    COLUMNS = {
        0: 'id',
        1: 'org_name',
        2: 'in_cnie',
        3: 'in_cace',
        4: 'in_un',
        5: 'founded_date',
        6: 'go_global_date',
        7: 'leaders',
        8: 'key_staff',
        9: 'capital_type',
        10: 'reg_location',
        11: 'reg_type',
        12: 'donation_pre',
        13: 'donation_pre_year',
        14: 'donation_post',
        15: 'donation_post_year',
        16: 'mission',
        17: 'org_structure',
        18: 'has_overseas_office',
        19: 'overseas_mission',
        20: 'overseas_projects',
        21: 'overseas_regions',
        22: 'overseas_services',
        23: 'service_mode',
        24: 'has_official_background',
        25: 'sources',
        26: 'disclosed_online',
        27: 'disclosed_continuous',
        28: 'go_out_level',
        29: 'logo_url'
    }

    print('\n3. Processing records...\n')

    success_count = 0
    error_count = 0
    updated_count = 0

    # Process each row (skip header row 1)
    for row_num in range(2, ws.max_row + 1):
        try:
            row = list(ws[row_num])

            # Get ID
            org_id = row[0].value
            if org_id is None:
                continue

            org_id = int(org_id)
            org_name = clean_value(row[1].value)

            if not org_name:
                continue

            # Build record dictionary
            record = {
                'id': org_id,
                'org_name': org_name,
                'in_cnie': parse_yes_no(row[2].value) if len(row) > 2 else None,
                'in_cace': parse_yes_no(row[3].value) if len(row) > 3 else None,
                'in_un': parse_yes_no(row[4].value) if len(row) > 4 else None,
                'founded_date': normalize_date(row[5].value) if len(row) > 5 else '——',
                'go_global_date': normalize_date(row[6].value) if len(row) > 6 else '——',
                'leaders': clean_value(row[7].value) if len(row) > 7 else None,
                'key_staff': clean_value(row[8].value) if len(row) > 8 else None,
                'capital_type': clean_value(row[9].value) if len(row) > 9 else None,
                'reg_location': clean_value(row[10].value) if len(row) > 10 else None,
                'reg_type': clean_value(row[11].value) if len(row) > 11 else None,
                'donation_pre': float(row[12].value) if (len(row) > 12 and row[12].value and row[12].value != '——') else None,
                'donation_pre_year': clean_value(row[13].value) if len(row) > 13 else None,
                'donation_post': float(row[14].value) if (len(row) > 14 and row[14].value and row[14].value != '——') else None,
                'donation_post_year': clean_value(row[15].value) if len(row) > 15 else None,
                'mission': clean_value(row[16].value) if len(row) > 16 else None,
                'org_structure': clean_value(row[17].value) if len(row) > 17 else None,
                'has_overseas_office': parse_yes_no(row[18].value) if len(row) > 18 else None,
                'overseas_mission': clean_value(row[19].value) if len(row) > 19 else None,
                'overseas_projects': clean_value(row[20].value) if len(row) > 20 else None,
                'overseas_regions': clean_value(row[21].value) if len(row) > 21 else None,
                'overseas_services': clean_value(row[22].value) if len(row) > 22 else None,
                'service_mode': clean_value(row[23].value) if len(row) > 23 else None,
                'has_official_background': parse_yes_no(row[24].value) if len(row) > 24 else None,
                'sources': clean_value(row[25].value) if len(row) > 25 else None,
                'disclosed_online': parse_yes_no(row[26].value) if len(row) > 26 else None,
                'disclosed_continuous': parse_yes_no(row[27].value) if len(row) > 27 else None,
                'go_out_level': clean_value(row[28].value) if len(row) > 28 else None,
                'logo_url': logo_backup.get(org_id, None)
            }

            # Build UPDATE SQL (to update existing records)
            update_fields = []
            for field in ['org_name', 'in_cnie', 'in_cace', 'in_un', 'founded_date',
                         'go_global_date', 'leaders', 'key_staff', 'capital_type',
                         'reg_location', 'reg_type', 'donation_pre', 'donation_pre_year',
                         'donation_post', 'donation_post_year', 'mission', 'org_structure',
                         'has_overseas_office', 'overseas_mission', 'overseas_projects',
                         'overseas_regions', 'overseas_services', 'service_mode',
                         'has_official_background', 'sources', 'disclosed_online',
                         'disclosed_continuous', 'go_out_level', 'logo_url']:

                val = record[field]
                if val is None:
                    update_fields.append(f"{field} = NULL")
                elif isinstance(val, (int, float)):
                    update_fields.append(f"{field} = {val}")
                else:
                    update_fields.append(f"{field} = {escape_sql_string(val)}")

            update_sql = f"UPDATE orgs SET {', '.join(update_fields)} WHERE id = {org_id};"

            # Execute update
            execute_sql(DB_NAME, update_sql)

            # Show progress every 50 records
            if org_id % 50 == 0:
                print(f'   ✓ ID {org_id}: {org_name[:40]}')

            updated_count += 1

        except Exception as e:
            print(f'   ✗ Row {row_num} (ID {org_id if "org_id" in locals() else "?"}): {str(e)[:80]}')
            error_count += 1

    print(f'\n' + '='*100)
    print(f'✅ Import Complete')
    print(f'   Updated: {updated_count} records')
    print(f'   Errors: {error_count} records')
    print('='*100)

if __name__ == '__main__':
    main()
